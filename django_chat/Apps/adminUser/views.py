from django.shortcuts import render
from django.conf import settings
from django.http import JsonResponse
from django.views.generic import View
from Apps.main.models import User, CountUser, Answer, HighProblem
from django.core import serializers
from django.contrib.auth.hashers import make_password
import openpyxl
import datetime
import os
import re


# 后台管理主页
class AdminUser(View):
    def get(self, request):
        user = serializers.serialize('python', User.objects.all())

        return JsonResponse({'user': user}, safe=False, json_dumps_params={'ensure_ascii': False})


# 删除用户
class DelUserView(View):
    def post(self, request, uid):
        try:
            User.objects.get(id=uid).delete()
            return JsonResponse({'res': 0, 'message': '删除成功'})
        except Exception:
            return JsonResponse({'res': 1, 'message': '没有这个用户'})


# 重置密码
class ResetPwdView(View):
    def get(self, request, uid):
        try:
            user = User.objects.get(id=uid)
        except Exception:
            return JsonResponse({'res': 1, 'message': '无效用户'})

        return JsonResponse({'res': 0, 'user': user})

    def post(self, request, uid):
        new_pwd = request.POST.get('newpwd')
        print(uid)

        pw = re.compile(r'^[a-zA-Z0-9]{6,20}$')
        if not pw.search(new_pwd):
            return JsonResponse({'res': 4, 'message': '密码格式错误'})

        try:
            # 更新的密码进行加密
            new_pwd = make_password(password=new_pwd, salt=None, hasher='pbkdf2_sha1')
            User.objects.filter(id=uid).update(password=new_pwd)
            return JsonResponse({'res': 0, 'message': '更新成功'})
        except Exception as e:
            print(e)
            return JsonResponse({'res': 1, 'message': '无效用户'})


# 禁止登陆
class BanUserView(View):
    def post(self, request, uid):
        try:
            user = User.objects.get(id=uid)
            user.is_active = 0
            user.save()
            return JsonResponse({'res': 0, 'message': '禁用成功'})
        except:
            return JsonResponse({'res': 1, 'message': '无效用户，禁用失败'})


# 应答文字管理
class AnswerView(View):
    def get(self, request):
        answer = serializers.serialize('python', Answer.objects.all())

        return JsonResponse({'answer': answer}, safe=False, json_dumps_params={'ensure_ascii': False})


class DelAnswerView(View):
    def post(self, request, aid):
        try:
            Answer.objects.get(id=aid).delete()
            return JsonResponse({'res': 0, 'message': '删除成功'})
        except Exception:
            return JsonResponse({'res': 1, 'message': '没有这个应答'})


class EditAnswerView(View):
    def get(self, request, aid):
        try:
            answer = Answer.objects.get(id=aid)
            return JsonResponse({'res': 0, 'answer': answer})
        except:
            return JsonResponse({'res': 1, 'message': '无效的数据'})

    def post(self, request, aid):
        solutions = request.POST.get('solutions')
        img = request.FILES.get('img')
        url = request.POST.get('url', '')
        try:
            answer = Answer.objects.get(id=aid)
            img_name = img.name

            if img:
                with open(f'/static/img/{img_name}', 'wb+') as f:
                    for chunks in img.chunks():
                        f.write(chunks)

            answer.solutions = solutions
            answer.img = img_name
            answer.url = url
            answer.save()
            return JsonResponse({'res': 0, 'answer': answer})
        except:
            return JsonResponse({'res': 1, 'message': '无效的数据'})


# 统计管理
class CountMsgView(View):
    def get(self, request):
        count_user = serializers.serialize('python', Answer.objects.all())
        high_count = serializers.serialize('python', HighProblem.objects.all().order_by('-ask_time'))
        return JsonResponse({'count_user': count_user, 'high_count': high_count}, safe=False, json_dumps_params={'ensure_ascii': False})


# 用户导入导出
class UserExcelView(View):
    # 导出
    def get(self, request):
        obj_user = User.objects.all().values()
        userlist = list(obj_user)
        excel_name = datetime.datetime.now().strftime('%Y-%m-%d') + '.xlsx'

        path = os.path.join(settings.MEDIA_ROOT, excel_name)
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'user'
        keys = userlist[0].keys()
        for index, item in enumerate(User.get_title_list()):
            sheet.cell(row=1, column=index + 1, value=item)

        for index, item in enumerate(userlist):
            for k, v in enumerate(keys):
                sheet.cell(row=index + 2, column=k + 1, value=str(item[v]))

        workbook.save(path)
        return JsonResponse({'res': 0, 'message': '导出成功'})

    # 导入
    def post(self, request):
        pass


# 统计导出
class CountExcelView(View):
    def post(self, request):
        obj_user = HighProblem.objects.all().values()
        highlist = list(obj_user)
        excel_name = 'high_proble' + datetime.datetime.now().strftime('%Y-%m-%d') + '.xlsx'

        path = os.path.join(settings.MEDIA_ROOT, excel_name)
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'high_proble'
        keys = highlist[0].keys()
        for index, item in enumerate(HighProblem.get_title_list()):
            sheet.cell(row=1, column=index + 1, value=item)

        for index, item in enumerate(highlist):
            for k, v in enumerate(keys):
                sheet.cell(row=index + 2, column=k + 1, value=str(item[v]))

        workbook.save(path)
        return JsonResponse({'res': 0, 'message': '导出成功'})
